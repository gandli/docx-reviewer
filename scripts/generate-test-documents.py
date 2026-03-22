from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
import xlwt


ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "src" / "tests" / "fixtures" / "documents"

TITLE = "办公设备升级采购需求说明书"

MARKDOWN_CONTENT = f"""# {TITLE}

## 一、项目背景

随着公司业务规模扩大和远程协作常态化，现有部分办公设备已连续使用超过 4 年，出现性能下降、电池续航缩短和硬件故障率升高等问题。

## 二、采购目标

1. 提升员工日常办公效率
2. 降低设备故障造成的业务中断风险
3. 统一 IT 资产管理口径

## 三、预算与付款

- 预算上限：人民币 225,000 元
- 付款建议：验收通过且发票齐备后，按约定节点分阶段支付

## 四、交付要求

| 项目 | 内容 |
| --- | --- |
| 交付时间 | 合同签订后 7 个工作日内 |
| 交付地点 | 苏州工业园区示例地址 |
| 保修要求 | 3 年原厂上门保修 |
"""

TXT_CONTENT = f"""{TITLE}

一、项目背景
随着公司业务规模扩大和远程协作常态化，现有部分办公设备已连续使用超过 4 年，出现性能下降、电池续航缩短和硬件故障率升高等问题。

二、采购目标
1. 提升员工日常办公效率
2. 降低设备故障造成的业务中断风险
3. 统一 IT 资产管理口径

三、预算与付款
预算上限为人民币 225,000 元。
建议在验收通过且发票齐备后，按约定节点分阶段支付。
"""

HTML_CONTENT = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <title>{TITLE}</title>
  <style>
    body {{
      font-family: "PingFang SC", "Microsoft YaHei", sans-serif;
      margin: 48px;
      color: #222;
      line-height: 1.7;
      font-size: 13px;
    }}
    h1, h2 {{
      color: #294e80;
      margin: 0 0 16px;
    }}
    h1 {{
      text-align: center;
      font-size: 24px;
      margin-bottom: 28px;
    }}
    h2 {{
      font-size: 18px;
      margin-top: 24px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      margin-top: 18px;
    }}
    th, td {{
      border: 1px solid #333;
      padding: 8px 10px;
      text-align: left;
      vertical-align: top;
    }}
    ul {{
      padding-left: 20px;
    }}
  </style>
</head>
<body>
  <h1>{TITLE}</h1>
  <h2>一、项目背景</h2>
  <p>随着公司业务规模扩大和远程协作常态化，现有部分办公设备已连续使用超过 4 年，出现性能下降、电池续航缩短和硬件故障率升高等问题。</p>

  <h2>二、采购目标</h2>
  <ul>
    <li>提升员工日常办公效率</li>
    <li>降低设备故障造成的业务中断风险</li>
    <li>统一 IT 资产管理口径</li>
  </ul>

  <h2>三、预算与付款</h2>
  <p>预算上限为人民币 225,000 元。建议在验收通过且发票齐备后，按约定节点分阶段支付。</p>

  <h2>四、交付要求</h2>
  <table>
    <tr><th>项目</th><th>内容</th></tr>
    <tr><td>交付时间</td><td>合同签订后 7 个工作日内</td></tr>
    <tr><td>交付地点</td><td>苏州工业园区示例地址</td></tr>
    <tr><td>保修要求</td><td>3 年原厂上门保修</td></tr>
  </table>
</body>
</html>
"""


def write_text_files() -> tuple[Path, Path, Path]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    md_path = OUT_DIR / "sample-procurement-spec.md"
    txt_path = OUT_DIR / "sample-procurement-spec.txt"
    html_path = OUT_DIR / "sample-procurement-spec.html"

    md_path.write_text(MARKDOWN_CONTENT, encoding="utf-8")
    txt_path.write_text(TXT_CONTENT, encoding="utf-8")
    html_path.write_text(HTML_CONTENT, encoding="utf-8")
    return md_path, txt_path, html_path


def run_textutil(html_path: Path, fmt: str) -> None:
    output_path = OUT_DIR / f"sample-procurement-spec.{fmt}"
    subprocess.run(
      [
        "textutil",
        "-convert",
        fmt,
        str(html_path),
        "-output",
        str(output_path),
      ],
      check=True,
    )


def build_pdf() -> None:
    pdf_path = OUT_DIR / "sample-procurement-spec.pdf"
    styles = getSampleStyleSheet()
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    title_style = ParagraphStyle(
      "FixtureTitle",
      parent=styles["Title"],
      fontName="STSong-Light",
      fontSize=22,
      leading=28,
      textColor=colors.HexColor("#294e80"),
      alignment=1,
    )
    heading_style = ParagraphStyle(
      "FixtureHeading",
      parent=styles["Heading2"],
      fontName="STSong-Light",
      fontSize=16,
      leading=22,
      textColor=colors.HexColor("#294e80"),
      spaceBefore=8,
      spaceAfter=8,
    )
    body_style = ParagraphStyle(
      "FixtureBody",
      parent=styles["BodyText"],
      fontName="STSong-Light",
      fontSize=11,
      leading=18,
      textColor=colors.HexColor("#222222"),
    )
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4, title=TITLE)

    story = [
      Paragraph(TITLE, title_style),
      Spacer(1, 12),
      Paragraph("一、项目背景", heading_style),
      Paragraph(
        "随着公司业务规模扩大和远程协作常态化，现有部分办公设备已连续使用超过 4 年，"
        "出现性能下降、电池续航缩短和硬件故障率升高等问题。",
        body_style,
      ),
      Spacer(1, 10),
      Paragraph("二、采购目标", heading_style),
      Paragraph("1. 提升员工日常办公效率", body_style),
      Paragraph("2. 降低设备故障造成的业务中断风险", body_style),
      Paragraph("3. 统一 IT 资产管理口径", body_style),
      Spacer(1, 10),
      Paragraph("三、预算与付款", heading_style),
      Paragraph("预算上限为人民币 225,000 元。", body_style),
      Paragraph("建议在验收通过且发票齐备后，按约定节点分阶段支付。", body_style),
      Spacer(1, 14),
    ]

    table = Table(
      [
        ["项目", "内容"],
        ["交付时间", "合同签订后 7 个工作日内"],
        ["交付地点", "苏州工业园区示例地址"],
        ["保修要求", "3 年原厂上门保修"],
      ],
      colWidths=[120, 320],
    )
    table.setStyle(
      TableStyle(
        [
          ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbe5f1")),
          ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
          ("GRID", (0, 0), (-1, -1), 0.8, colors.HexColor("#333333")),
          ("VALIGN", (0, 0), (-1, -1), "TOP"),
          ("PADDING", (0, 0), (-1, -1), 6),
          ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
          ("FONTSIZE", (0, 0), (-1, -1), 11),
        ]
      )
    )
    story.append(table)

    doc.build(story)


def build_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "采购清单"
    ws.append(["序号", "项目", "数量", "单价", "小计"])
    rows = [
      [1, "商务笔记本电脑", 25, 8200, "=C2*D2"],
      [2, "原装设备包", 25, 300, "=C3*D3"],
      [3, "软件预装与配置服务", 1, 1800, "=C4*D4"],
    ]
    for row in rows:
        ws.append(row)
    ws["E5"] = "=SUM(E2:E4)"
    ws["A7"] = "付款建议"
    ws["B7"] = "验收通过且发票齐备后，按约定节点分阶段支付"

    info = wb.create_sheet("供应商要求")
    info.append(["项目", "要求"])
    info.append(["交付时间", "合同签订后 7 个工作日内"])
    info.append(["保修要求", "3 年原厂上门保修"])

    wb.save(OUT_DIR / "sample-procurement-spec.xlsx")


def build_xls() -> None:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("采购清单")
    headers = ["序号", "项目", "数量", "单价", "小计"]
    for col, value in enumerate(headers):
        sheet.write(0, col, value)

    rows = [
      [1, "商务笔记本电脑", 25, 8200, 205000],
      [2, "原装设备包", 25, 300, 7500],
      [3, "软件预装与配置服务", 1, 1800, 1800],
    ]
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row):
            sheet.write(row_index, col_index, value)

    sheet.write(5, 0, "付款建议")
    sheet.write(5, 1, "验收通过且发票齐备后，按约定节点分阶段支付")
    workbook.save(str(OUT_DIR / "sample-procurement-spec.xls"))


def write_manifest() -> None:
    manifest = OUT_DIR / "README.md"
    manifest.write_text(
        """# 测试文件样例

这组文件用于导入、原样预览和格式兼容测试，内容主题保持一致，便于不同格式之间对照。

## 重新生成

```bash
uv run --with openpyxl --with reportlab --with xlwt python scripts/generate-test-documents.py
```

## 文件列表

- `sample-procurement-spec.doc`
- `sample-procurement-spec.docx`
- `sample-procurement-spec.pdf`
- `sample-procurement-spec.md`
- `sample-procurement-spec.txt`
- `sample-procurement-spec.xls`
- `sample-procurement-spec.xlsx`

## 内容主题

统一使用“办公设备升级采购需求说明书”这一主题，包含：

- 项目背景
- 采购目标
- 预算与付款建议
- 简单表格
""",
        encoding="utf-8",
    )


def main() -> int:
    _, _, html_path = write_text_files()
    run_textutil(html_path, "docx")
    run_textutil(html_path, "doc")
    build_pdf()
    build_xlsx()
    build_xls()
    write_manifest()
    print(f"generated test documents in: {OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
